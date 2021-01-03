# library to open excel files
import openpyxl as px
import numpy as np
import warnings
import nltk
# loading the workbook which containes the positive words
# I have refererd the site http://positivewordsresearch.com/list-of-positive-words/ to write about 614 positive words
W = px.load_workbook('list_of_postive_words.xlsx')
# fetch the file using the sheet name
p = W.get_sheet_by_name(name = 'Sheet1')
# empty list which is used to append the elements
a=[]
# p.iter_rows() will refer to the cell in the sheet which contains the value
for row in p.iter_rows():
    for k in row:
# value will be fetched using internal_value and words if capital are convereted to lower case using lower function
        c = k.internal_value.lower()
# every word from the excel sheet is appended to the list
        a.append(c)
print("Please see the list of positive words:")
print(a)
x = px.load_workbook('list_of_negative_words.xlsx')
p = x.get_sheet_by_name(name = 'Sheet1')
b=[]
for row in p.iter_rows():
    for k in row:
        c = k.internal_value.lower()
        b.append(c)
print("Please see the list of negative words:")
print(b)
# this is the main function where both kist of positive and negative words will be used.
# please install nltk for such projects. I have used anaconda to install such projects.
def sentimental_analysis(text):
# first of all text is xonverted into sentence    
    text_sentence = nltk.sent_tokenize(text)
    for sentence in text_sentence:
        negative_count = 0
        positive_count= 0
        # second apporach will be converting the sentence into words
        words_count_sentence = nltk.word_tokenize(sentence)
        for words in words_count_sentence:
            print(words)
            for items in a:
                # comparing the words of the sentence with the list of positive words in the list
                 if (words == items):
                    positive_count += 1

            for items in b:
                 # comparing the words of the sentence with the list of positive words in the list
                if (words == items):
                    negative_count += 1
        print("Please see positive count:", positive_count)
        print("Please see negative count:",negative_count)
# please see my conditions on the positive and negative counts. You can change accordingly.
        if positive_count >0 and negative_count ==0:
            print("Written text is postive and person is writing optimistic")
        elif negative_count>0 and positive_count == 0:
            print("Written text is negative and person is writing pessimistic")
        elif positive_count>negative_count and negative_count!=0:
            print("Person is writing more positive and less negative")
        elif negative_count>positive_count and positive_count!=0:
            print("Person is writing more negative and less positve")
        else:
            print("Person is writing neutral")
# user can input the text from here.
text = input("Please enter your sentence to check its state:")
# text is converted into lower case for comparing each words with words in the list
text = text.lower()
output = sentimental_analysis(text)
